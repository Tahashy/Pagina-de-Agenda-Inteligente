from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import pandas as pd
from pandas.api.types import is_datetime64tz_dtype, is_datetime64_dtype


# --------------------------------------------------------
# Normaliza fechas: convierte columnas datetime con tz a sin tz
# --------------------------------------------------------
def _ensure_naive_datetimes(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return df

    for col in df.columns:
        ser = df[col]

        try:
            # Caso 1: dtype datetime con zona horaria
            if is_datetime64tz_dtype(ser):
                df[col] = ser.dt.tz_convert("UTC").dt.tz_localize(None)
                continue

            # Caso 2: dtype datetime sin zona
            if is_datetime64_dtype(ser):
                df[col] = pd.to_datetime(ser, errors="coerce")
                continue

            # Caso 3: columna object con strings que parecen fechas
            if ser.dtype == object:
                parsed = pd.to_datetime(ser, errors="coerce")

                # Si el parseo funcionó, reemplazar
                if parsed.notna().sum() > 0:
                    # Si trae TZ → normalizar
                    if is_datetime64tz_dtype(parsed):
                        parsed = parsed.dt.tz_convert("UTC").dt.tz_localize(None)

                    df[col] = parsed

        except Exception:
            # Ignorar errores sin romper el proceso (seguro)
            pass

    return df


# --------------------------------------------------------
# Reporte mensual consolidado (varias hojas)
# --------------------------------------------------------
def generar_reporte_mensual_excel(nombre_archivo, resumen: dict,
                                  df_inc: pd.DataFrame,
                                  df_epp: pd.DataFrame,
                                  df_cap: pd.DataFrame,
                                  by_area: pd.DataFrame = None):
    with pd.ExcelWriter(nombre_archivo, engine="openpyxl") as writer:

        # Resumen
        df_res = pd.DataFrame(list(resumen.items()), columns=["Métrica", "Valor"])
        df_res.to_excel(writer, sheet_name="Resumen", index=False)

        # Incidentes
        if df_inc is None or df_inc.empty:
            pd.DataFrame([{"mensaje": "No hay incidentes"}]).to_excel(
                writer, sheet_name="Incidentes", index=False)
        else:
            df_inc = _ensure_naive_datetimes(df_inc.copy())
            df_inc.to_excel(writer, sheet_name="Incidentes", index=False)

        # EPP
        if df_epp is None or df_epp.empty:
            pd.DataFrame([{"mensaje": "No hay registros EPP"}]).to_excel(
                writer, sheet_name="EPP", index=False)
        else:
            df_epp = _ensure_naive_datetimes(df_epp.copy())
            df_epp.to_excel(writer, sheet_name="EPP", index=False)

        # Capacitaciones
        if df_cap is None or df_cap.empty:
            pd.DataFrame([{"mensaje": "No hay capacitaciones"}]).to_excel(
                writer, sheet_name="Capacitaciones", index=False)
        else:
            df_cap = _ensure_naive_datetimes(df_cap.copy())
            df_cap.to_excel(writer, sheet_name="Capacitaciones", index=False)

        # Incidentes por área (agrupado)
        if by_area is None or by_area.empty:
            pd.DataFrame([{"mensaje": "Sin datos por área"}]).to_excel(
                writer, sheet_name="Incidentes_por_area", index=False)
        else:
            by_area = _ensure_naive_datetimes(by_area.copy())
            by_area.to_excel(writer, sheet_name="Incidentes_por_area", index=False)

    return nombre_archivo


# --------------------------------------------------------
# Exportación simple a Excel (una sola hoja)
# --------------------------------------------------------
def exportar_excel(nombre_archivo, df: pd.DataFrame):
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte SST"

    # Estilos
    header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    header_font = Font(bold=True)
    align_center = Alignment(horizontal="center", vertical="center")

    # Encabezados
    for col, columna in enumerate(df.columns, start=1):
        c = ws.cell(row=1, column=col, value=columna)
        c.fill = header_fill
        c.font = header_font
        c.alignment = align_center

    # Filas
    for row_idx, fila in df.iterrows():
        for col_idx, valor in enumerate(fila, start=1):
            ws.cell(row=row_idx + 2, column=col_idx, value=str(valor))

    # Ajustar ancho columnas
    for col in ws.columns:
        max_len = max(len(str(cell.value)) for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + 4

    wb.save(nombre_archivo)
    return nombre_archivo
